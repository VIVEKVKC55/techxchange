from django.contrib import admin, messages
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404, redirect
from django.core.mail import send_mail
from django.utils.timezone import now
from datetime import timedelta
from django.utils.html import format_html
from django.urls import path, reverse
from .models import Subscription, UserProfile, SubscriptionPlan, PlanType, SubscriptionDuration, PaymentRecord
from django.conf import settings
from django.contrib.auth.admin import UserAdmin as DefaultUserAdmin
from django.utils.safestring import mark_safe
import secrets
import string
from django.http import HttpResponse
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from django.core.mail import EmailMessage
import logging
logger = logging.getLogger(__name__)

User = get_user_model()

admin.site.index_title = "Techexchange Dashboard"      # Dashboard subtitle

@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = ("user", "phone_number", "is_verified", "created_at")
    search_fields = ("user__username", "phone_number")
    list_filter = ("is_verified", "created_at")


class CustomUserAdmin(DefaultUserAdmin):
    list_display = (
        "email", "username", "user_category", "phone_number",
        "address", "profile_picture_preview", "is_verified",
        "date_joined", "is_active"
    )
    list_filter = ("subscription__plan", "is_active", "profile__is_verified")
    search_fields = ("email", "username", "profile__phone_number")
    actions = ["approve_users", "download_filtered_user_data"]  # Add bulk actions

    def user_category(self, obj):
        if hasattr(obj, "subscription"):
            if obj.subscription.plan.name == "Premium (Plan A)":
                return "Premium (Plan A)"
            elif obj.subscription.plan.name == "Premium (Plan B)":
                return "Premium (Plan B)"
            return "Paid User"
        return "Free User"

    user_category.short_description = "User Type"

    def phone_number(self, obj):
        return obj.profile.phone_number if hasattr(obj, "profile") and obj.profile else "N/A"

    phone_number.short_description = "Phone"

    def address(self, obj):
        return obj.profile.address if hasattr(obj, "profile") and obj.profile else "N/A"

    address.short_description = "Address"

    def is_verified(self, obj):
        return obj.profile.is_verified if hasattr(obj, "profile") and obj.profile else False

    is_verified.short_description = "Verified"
    is_verified.boolean = True

    def profile_picture_preview(self, obj):
        if hasattr(obj, "profile") and obj.profile and obj.profile.profile_picture:
            return format_html(f'<img src="{obj.profile.profile_picture.url}" width="40" height="40" style="border-radius:50%;">')
        return "No Image"

    profile_picture_preview.short_description = "Profile Pic"

    @admin.action(description="Approve selected users and send login details")
    def approve_users(self, request, queryset):
        for user in queryset:
            if not user.is_active:
                characters = string.ascii_letters + string.digits + string.punctuation
                default_password = ''.join(secrets.choice(characters) for _ in range(6))
                user.set_password(default_password)
                user.is_active = True
                user.profile.set_password(default_password)
                user.save()

                subject = "Your Account is Approved!"
                message = f"""
                Dear {user.first_name},

                Your account has been approved by the admin.

                You can now log in with the following details:

                Email: {user.email}
                Password: {default_password}

                Please change your password after logging in.

                Thank you!
                """
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email])

        self.message_user(request, "Selected users have been approved and notified via email.")

    @admin.action(description="Download filtered user data as Excel")
    def download_filtered_user_data(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter

        changelist = self.get_changelist_instance(request)
        filtered_queryset = changelist.get_queryset(request)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Filtered Users'

        headers = [
            "Full Name", "Username", "Email", "User Plan Category", "Phone", "Address",
            "Verified", "Active", "Date Joined"
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

        for row_num, user in enumerate(filtered_queryset, 2):
            profile = getattr(user, 'profile', None)
            sheet[f"A{row_num}"] = user.get_full_name() or 'N/A'
            sheet[f"B{row_num}"] = user.username
            sheet[f"C{row_num}"] = user.email
            sheet[f"D{row_num}"] = self.user_category(user)
            sheet[f"E{row_num}"] = getattr(profile, 'phone_number', 'N/A')
            sheet[f"F{row_num}"] = getattr(profile, 'address', 'N/A')
            sheet[f"G{row_num}"] = "Yes" if getattr(profile, 'is_verified', False) else "No"
            sheet[f"H{row_num}"] = "Yes" if user.is_active else "No"
            sheet[f"I{row_num}"] = user.date_joined.strftime('%Y-%m-%d %H:%M')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=filtered_user_data.xlsx'
        workbook.save(response)
        return response

    def change_view(self, request, object_id, form_url='', extra_context=None):
        user = get_object_or_404(User, pk=object_id)
        decrypted_password = None
        if hasattr(user, "profile") and user.profile and user.profile.encrypted_password:
            try:
                decrypted_password = user.profile.get_password()
            except Exception:
                decrypted_password = "[Could not decrypt password]"

        extra_context = extra_context or {}
        extra_context['decrypted_password'] = decrypted_password
        return super().change_view(request, object_id, form_url, extra_context=extra_context)


# Unregister the default User model if already registered
admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)


class SubscriptionAdmin(admin.ModelAdmin):
    list_display = (
        "user", "approve_button", "amount_paid", "plan", "duration_days", "is_approved",
        "remaining_days_display", "pending_plan", "pending_duration", "start_date",
        "end_date", "reject_button"
    )
    list_filter = ("plan", "is_approved")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('approve/<int:subscription_id>/', self.admin_site.admin_view(self.approve_subscription), name="approve_subscription"),
            path('reject/<int:subscription_id>/', self.admin_site.admin_view(self.reject_subscription), name="reject_subscription"),
        ]
        return custom_urls + urls

    def approve_subscription(self, request, subscription_id):
        subscription = get_object_or_404(Subscription, id=subscription_id)

        if subscription.amount_paid <= 0:
            messages.error(request, "Cannot approve. Payment not received yet.")
            return redirect('/admin/customer/subscription/')

        if subscription.pending_plan and subscription.pending_duration:
            plan_type = subscription.pending_plan
            duration = subscription.pending_duration

            try:
                subscription_plan = SubscriptionPlan.objects.get(plan_type=plan_type, duration_days=duration)
            except SubscriptionPlan.DoesNotExist:
                messages.error(request, "Invalid plan-duration combination.")
                return redirect('/admin/customer/subscription/')

            try:
                # Update subscription
                subscription.plan = plan_type
                subscription.duration_days = duration
                subscription.start_date = now()
                subscription.end_date = now() + timedelta(days=duration.duration_days)
                subscription.is_approved = True
                subscription.pending_plan = None
                subscription.pending_duration = None
                subscription.save()

                PaymentRecord.objects.create(
                    user=subscription.user,
                    subscription=subscription,
                    sub_amount=subscription.amount_paid,
                    payment_method="Cash",
                    paid_amount=subscription.amount_paid,
                    sub_remaining_amount=0,  # Assuming full payment
                    payment_date=now(),
                )

                # Generate styled PDF invoice
                buffer = BytesIO()
                p = canvas.Canvas(buffer, pagesize=letter)
                width, height = letter

                p.roundRect(0.5*inch, height - 4*inch, 7.5*inch, 3.5*inch, 10, stroke=1, fill=0)

                # Header
                p.setFont("Helvetica-Bold", 14)
                p.setFillColorRGB(0, 0.2, 0.6)
                p.rect(0.5*inch, height - 0.7*inch, 7.5*inch, 0.4*inch, fill=1)
                p.setFillColorRGB(1, 1, 1)
                p.drawString(0.6*inch, height - 0.55*inch, "CASH RECEIPT")

                # Content
                p.setFillColorRGB(0, 0, 0)
                p.setFont("Helvetica", 11)
                line_start = height - 1.1*inch
                line_gap = 0.3*inch

                p.drawString(0.6*inch, line_start, f"RECEIVED FROM: {subscription.user.username}")
                p.drawString(4.5*inch, line_start, f"DATE: {subscription.start_date.strftime('%Y-%m-%d')}")

                p.drawString(0.6*inch, line_start - line_gap, f"ADDRESS: {subscription.user.profile.address}")
                p.drawString(0.6*inch, line_start - 2*line_gap, f"DOLLAR (AED): {subscription.amount_paid:.2f}       FOR: {subscription.plan.name}")

                # Summary Table
                table_top = line_start - 3.5*line_gap
                p.drawString(0.6*inch, table_top, f"TOTAL DUE: AED {subscription.amount_paid:.2f}")
                p.drawString(0.6*inch, table_top - line_gap, f"PAID AMOUNT: AED {subscription.amount_paid:.2f}")
                p.drawString(0.6*inch, table_top - 2*line_gap, "DUE BALANCE: AED 0.00")

                p.drawString(3.5*inch, table_top, "CHECK")
                p.drawString(3.5*inch, table_top - line_gap, "CASH ✓")
                p.drawString(3.5*inch, table_top - 2*line_gap, "MONEY ORDER")

                p.drawString(0.6*inch, table_top - 3*line_gap, f"BY: Admin")

                p.showPage()
                p.save()

                buffer.seek(0)
                pdf_file = buffer.getvalue()
                buffer.close()

                # Send email with PDF
                email = EmailMessage(
                    "Subscription Approved",
                    f"Dear {subscription.user.username},\n\n"
                    f"Your subscription has been approved.\n\n"
                    f"Plan: {plan_type.name}\n"
                    f"Duration: {duration.duration_days} days\n"
                    f"Amount Paid: AED {subscription.amount_paid:.2f}\n\n"
                    "Please find the invoice attached.\n\nBest Regards,\nSupport Team",
                    "support@example.com",
                    [subscription.user.email]
                )
                email.attach("Invoice.pdf", pdf_file, "application/pdf")
                email.send(fail_silently=True)

                messages.success(request, f"Subscription for {subscription.user.username} approved and invoice emailed.")

            except Exception as e:
                logger.error(f"Error processing subscription approval: {e}")
                messages.error(request, "Error processing subscription approval. Please try again.")

        return redirect('/admin/customer/subscription/')

    def reject_subscription(self, request, subscription_id):
        subscription = get_object_or_404(Subscription, id=subscription_id)
        if subscription.pending_plan:
            subscription.pending_plan = None
            subscription.pending_duration = None
            subscription.is_approved = False
            subscription.save()

            send_mail(
                "Subscription Upgrade Rejected",
                f"Dear {subscription.user.username},\n\n"
                "Your subscription upgrade request was rejected.\n\n"
                "Best Regards,\nSupport Team",
                "support@example.com",
                [subscription.user.email],
                fail_silently=True,
            )

            messages.warning(request, f"Subscription for {subscription.user.username} rejected.")
        return redirect('/admin/customer/subscription/')

    def approve_button(self, obj):
        if obj.pending_plan:
            url = reverse('admin:approve_subscription', args=[obj.id])  # ✅ Ensure correct URL
            return mark_safe(
                f'<button class="button approve-btn" data-url="{url}" data-id="{obj.id}" data-amount="{obj.amount_paid}" style="background: green; color: white; padding: 5px 10px; border-radius: 4px;">Approve</button>'
            )
        return "No pending request"

    def reject_button(self, obj):
        if obj.pending_plan:
            url = reverse('admin:reject_subscription', args=[obj.id])  # ✅ Ensure correct URL
            return mark_safe(
                f'<button class="button reject-btn" data-url="{url}" data-id="{obj.id}" style="background: red; color: white; padding: 5px 10px; border-radius: 4px;">Reject</button>'
            )
        return "No pending request"

    def remaining_days_display(self, obj):
        return obj.remaining_days()

    remaining_days_display.short_description = "Remaining Days"
    approve_button.short_description = "Approve"
    reject_button.short_description = "Reject"

    class Media:
        js = ("admin/js/sweetalert2.min.js", "admin/js/subscription_actions.js")  # ✅ Load JavaScript for SweetAlert

admin.site.register(Subscription, SubscriptionAdmin)


@admin.register(PlanType)
class PlanTypeAdmin(admin.ModelAdmin):
    """Admin configuration for subscription plan types."""
    list_display = ("name", "max_products_per_day", "base_slots", "max_product_views_per_day")
    search_fields = ("name",)
    ordering = ("name",)


@admin.register(SubscriptionDuration)
class SubscriptionDurationAdmin(admin.ModelAdmin):
    """Admin configuration for subscription durations."""
    list_display = ("duration_days",)
    ordering = ("duration_days",)


@admin.register(SubscriptionPlan)
class SubscriptionPlanAdmin(admin.ModelAdmin):
    """Admin configuration for subscription plans (mapping plan type & duration)."""
    list_display = ("plan_type", "duration_days", "price")
    list_filter = ("plan_type", "duration_days")
    search_fields = ("plan_type__name",)
    ordering = ("plan_type", "duration_days")

from django.contrib import admin
from django.urls import path
from django.template.response import TemplateResponse
from django.utils.html import format_html
from django.core.mail import EmailMessage
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings

from .models import PaymentRecord
from django.contrib.auth.models import User
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from io import BytesIO
from django.core.mail import EmailMessage
from django.conf import settings
import datetime

class PaymentRecordAdmin(admin.ModelAdmin):
    list_display = ('user', 'subscription', 'paid_amount', 'payment_date', 'send_invoice_button')
    search_fields = ['user__username']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('send-invoice/<int:pk>/', self.admin_site.admin_view(self.send_invoice_view), name='send-invoice'),
        ]
        return custom_urls + urls

    def send_invoice_button(self, obj):
        return format_html(
            '<button class="button" onclick="openInvoicePopup({})">Send Invoice</button>', obj.id
        )
    send_invoice_button.short_description = 'Send Invoice'

    def send_invoice_view(self, request, pk):
        payment = get_object_or_404(PaymentRecord, pk=pk)

        if request.method == "POST":
            user_email = request.POST.get('email')
            sub_amount = request.POST.get('sub_amount')
            paid_amount = request.POST.get('paid_amount')
            remaining_amount = request.POST.get('sub_remaining_amount')
            payment_method = request.POST.get('payment_method')

            # Create PDF
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=LETTER)
            width, height = LETTER

            # Draw outer border
            p.setLineWidth(1)
            p.roundRect(50, 500, 500, 220, 10)

            # Header
            p.setFillColor(colors.darkblue)
            p.setStrokeColor(colors.black)
            p.rect(50, 700, 500, 20, fill=1)
            p.setFillColor(colors.white)
            p.setFont("Helvetica-Bold", 12)
            p.drawString(55, 705, "CASH RECEIPTS")

            p.setFillColor(colors.black)
            p.setFont("Helvetica", 10)

            # Received From, Date
            p.drawString(60, 675, "RECEIVED FROM:")
            p.drawString(300, 675, "DATE:")
            p.line(150, 673, 280, 673)
            p.line(340, 673, 530, 673)

            # Address
            p.drawString(60, 655, "ADDRESS:")
            p.line(130, 653, 530, 653)

            # Dollar and For
            p.drawString(60, 630, "DOLLAR ($):")
            p.drawString(300, 630, "FOR:")
            p.line(130, 628, 250, 628)
            p.line(330, 628, 530, 628)

            # Total Due, Paid, Balance
            p.drawString(60, 600, "TOTAL DUE:")
            p.line(130, 598, 250, 598)
            p.drawString(60, 580, "PAID AMOUNT:")
            p.line(130, 578, 250, 578)
            p.drawString(60, 560, "DUE BALANCE:")
            p.line(130, 558, 250, 558)

            # Payment methods
            p.drawString(300, 600, "CHECK:")
            p.line(360, 598, 530, 598)
            p.drawString(300, 580, "CASH:")
            p.line(360, 578, 530, 578)
            p.drawString(300, 560, "MONEY ORDER:")
            p.line(390, 558, 530, 558)

            # BY
            p.drawString(60, 530, "BY:")
            p.line(90, 528, 530, 528)

            # Fill in values
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            p.drawString(340, 675, today)
            p.drawString(130, 628, f"${sub_amount}")
            p.drawString(330, 628, "Subscription Fee")
            p.drawString(130, 598, f"${sub_amount}")
            p.drawString(130, 578, f"${paid_amount}")
            p.drawString(130, 558, f"${remaining_amount}")
            p.drawString(360, 578, payment_method)

            p.showPage()
            p.save()
            buffer.seek(0)

            email = EmailMessage(
                'Your Invoice',
                'Attached is your invoice.',
                settings.DEFAULT_FROM_EMAIL,
                [user_email],
            )
            email.attach('invoice.pdf', buffer.read(), 'application/pdf')
            email.send()

            buffer.seek(0)
            email = EmailMessage(
                'Your Subscription Invoice',
                'Please find attached your invoice.',
                settings.DEFAULT_FROM_EMAIL,
                [user_email],
            )
            email.attach('invoice.pdf', buffer.read(), 'application/pdf')
            email.send()

            return JsonResponse({'status': 'success'})

        return TemplateResponse(request, 'admin/send_invoice_popup.html', {'payment': payment})

admin.site.register(PaymentRecord, PaymentRecordAdmin)
